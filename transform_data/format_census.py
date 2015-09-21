__author__ = 'Juan Trelles Trabucco'
__email__ = "juantrelles@gmail.com"

'''
    Formats the data from the "Single year of age population estimates April 1, 2010 to July, 2014"
    to a list of INTERVAL years range. Filters of GENDER or the 999 years could be added but it is
    preferred to use Excel filters and the simply export the data to CSV.
    The formatted values are added as a new worksheet.
'''

from openpyxl import load_workbook

def gen_ranged_ws(file_name, interval):
    wb = load_workbook(file_name)
    ws1 = wb.worksheets[0]
    ws2 = wb.create_sheet()

    for i in range(1, 14):
        ws2.cell(row=1, column=i).value = "'" + ws1.cell(row=1, column=i).value + "'"

    curr_row = 2  # The first line is reserved for table headers
    age_init = -1
    age_end = -1
    sum_base2010 = 0
    sum_est2010 = 0
    sum_est2011 = 0
    sum_est2012 = 0
    sum_est2013 = 0
    sum_est2014 = 0

    for row in range(2, ws1.get_highest_row() + 1):
        # Read values in the selected row
        sumlev = ws1['A' + str(row)].value
        region = ws1['B' + str(row)].value
        division = ws1['C' + str(row)].value
        state = ws1['D' + str(row)].value
        name = ws1['E' + str(row)].value
        sex = ws1['F' + str(row)].value
        age = ws1['G' + str(row)].value
        base2010 = ws1['H' + str(row)].value
        est2010 = ws1['I' + str(row)].value
        est2011 = ws1['J' + str(row)].value
        est2012 = ws1['K' + str(row)].value
        est2013 = ws1['L' + str(row)].value
        est2014 = ws1['M' + str(row)].value

        if age == 999:
            # Write accumulated values in memory. Considers special case when accumulated rows
            # are less than age_end but the row gets to 999
            if sum_base2010 > 0:
                ranged_age = str(age_init) + " - " + str(age_end)
                write_row(ws2, curr_row, sumlev, region, division, state, name, sex, ranged_age,
                          sum_base2010, sum_base2010, sum_est2011, sum_est2012, sum_est2013, sum_est2014)
                curr_row += 1

            # write accumulated value for state or country
            write_row(ws2, curr_row, sumlev, region, division, state, name, sex, age,
                      base2010, est2010, est2011, est2012, est2013, est2014)
            curr_row += 1

            age_init, age_end, sum_base2010, sum_est2010, sum_est2011, sum_est2012, sum_est2013, sum_est2014 \
                = reset_values(age_init, age_end, base2010, est2010, est2011, est2012, est2013, est2014)
        else:
            if age_init == -1:
                # If starting to read or first element in a new state
                age_init, age_end = age, age + interval - 1

            sum_base2010 += ws1['H' + str(row)].value
            sum_est2010 += ws1['I' + str(row)].value
            sum_est2011 += ws1['J' + str(row)].value
            sum_est2012 += ws1['K' + str(row)].value
            sum_est2013 += ws1['L' + str(row)].value
            sum_est2014 += ws1['M' + str(row)].value

            if age == age_end:
                # If the sum has accumulated values for group_count elements
                ranged_age = str(age_init) + " - " + str(age_end)
                write_row(ws2, curr_row, sumlev, region, division, state, name, sex, ranged_age,
                          sum_base2010, sum_est2010, sum_est2011, sum_est2012, sum_est2013, sum_est2014)
                curr_row += 1

                age_init, age_end, sum_base2010, sum_est2010, sum_est2011, sum_est2012, sum_est2013, sum_est2014 \
                    = reset_values(age_init, age_end, base2010, est2010, est2011, est2012, est2013, est2014)

    wb.save(file_name)

def write_row(ws, row, sumlev, region, division, state, name, sex, age,
              base2010, est2010, est2011, est2012, est2013, est2014):
    ws['A' + str(row)].value = sumlev
    ws['B' + str(row)].value = region
    ws['C' + str(row)].value = division
    ws['D' + str(row)].value = state
    ws['E' + str(row)].value = name
    ws['F' + str(row)].value = sex
    ws['G' + str(row)].value = age
    ws['H' + str(row)].value = base2010
    ws['I' + str(row)].value = est2010
    ws['J' + str(row)].value = est2011
    ws['K' + str(row)].value = est2012
    ws['L' + str(row)].value = est2013
    ws['M' + str(row)].value = est2014

def reset_values(age_init, age_end, sum_base2010, sum_est2010, sum_est2011, sum_est2012, sum_est2013, sum_est2014):
    age_init = -1
    age_end = -1
    sum_base2010 = 0
    sum_est2010 = 0
    sum_est2011 = 0
    sum_est2012 = 0
    sum_est2013 = 0
    sum_est2014 = 0
    return age_init, age_end, sum_base2010, sum_est2010, sum_est2011, sum_est2012, sum_est2013, sum_est2014

# Execute program
gen_ranged_ws("SC-EST2014-AGESEX-CIV.xlsx", 10)
